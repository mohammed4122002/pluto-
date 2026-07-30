"""Data-source connectors for the import wizard. Each connector's job is
simply "give me a header row + data rows" (or, for Postgres, "give me a
list of tables/columns to pick from") — everything downstream (matching,
dry-run, execute) is a single generic pipeline regardless of source."""

import csv
import io
import json
import time
from datetime import datetime, timezone

import httpx
import jwt
from openpyxl import load_workbook

from app.core.security import decrypt_secret

MAX_HEADER_SCAN_ROWS = 10


def _pick_header_row(raw_rows: list[list]) -> int:
    """Some exports have title/merged rows above the real header — pick the
    row (within the first few) with the most non-empty text-like cells."""
    best_idx, best_count = 0, -1
    for idx, row in enumerate(raw_rows[:MAX_HEADER_SCAN_ROWS]):
        count = sum(1 for v in row if v is not None and str(v).strip() != "")
        if count > best_count:
            best_idx, best_count = idx, count
    return best_idx


def _rows_from_grid(grid: list[list]) -> tuple[list[str], list[dict]]:
    if not grid:
        return [], []
    header_idx = _pick_header_row(grid)
    headers = [str(h).strip() if h is not None else "" for h in grid[header_idx]]
    rows = []
    for values in grid[header_idx + 1 :]:
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        rows.append({headers[i]: values[i] for i in range(len(headers)) if i < len(values) and headers[i]})
    return [h for h in headers if h], rows


def read_file(content: bytes, filename: str) -> tuple[list[str], list[dict]]:
    if filename.lower().endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        grid = [list(row) for row in sheet.iter_rows(values_only=True)]
        return _rows_from_grid(grid)

    text = None
    for encoding in ("utf-8-sig", "cp1256", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = content.decode("utf-8", errors="replace")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    grid = list(reader)
    return _rows_from_grid(grid)


def list_excel_sheets(content: bytes) -> list[str]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return workbook.sheetnames


def read_excel_sheet(content: bytes, sheet_name: str) -> tuple[list[str], list[dict]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    grid = [list(row) for row in sheet.iter_rows(values_only=True)]
    return _rows_from_grid(grid)


# ---------------------------------------------------------------------------
# Google Sheets (Service Account, no OAuth consent flow needed) — the sheet
# must be shared with the service account's client_email as a Viewer.

_GOOGLE_TOKEN_CACHE: dict[str, tuple[str, float]] = {}


def _google_access_token(service_account_json: dict) -> str:
    client_email = service_account_json["client_email"]
    cached = _GOOGLE_TOKEN_CACHE.get(client_email)
    if cached and cached[1] > time.time() + 30:
        return cached[0]

    now = int(time.time())
    payload = {
        "iss": client_email,
        "scope": "https://www.googleapis.com/auth/spreadsheets.readonly",
        "aud": "https://oauth2.googleapis.com/token",
        "iat": now,
        "exp": now + 3600,
    }
    assertion = jwt.encode(payload, service_account_json["private_key"], algorithm="RS256")
    resp = httpx.post(
        "https://oauth2.googleapis.com/token",
        data={"grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer", "assertion": assertion},
        timeout=15,
    )
    resp.raise_for_status()
    body = resp.json()
    _GOOGLE_TOKEN_CACHE[client_email] = (body["access_token"], time.time() + body.get("expires_in", 3600))
    return body["access_token"]


def google_sheets_verify(service_account_json: dict, spreadsheet_id: str) -> dict:
    token = _google_access_token(service_account_json)
    resp = httpx.get(
        f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}",
        params={"fields": "properties.title,sheets.properties.title"},
        headers={"Authorization": f"Bearer {token}"},
        timeout=15,
    )
    if resp.status_code != 200:
        detail = resp.json().get("error", {}).get("message", "تعذر الوصول لملف الشيت")
        raise ValueError(f"{detail} — تأكد من مشاركة الشيت مع {service_account_json.get('client_email')} كمُشاهد")
    body = resp.json()
    return {
        "title": body["properties"]["title"],
        "tabs": [s["properties"]["title"] for s in body.get("sheets", [])],
    }


def google_sheets_read(service_account_json: dict, spreadsheet_id: str, tab: str) -> tuple[list[str], list[dict]]:
    token = _google_access_token(service_account_json)
    resp = httpx.get(
        f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{tab}",
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
    )
    resp.raise_for_status()
    values = resp.json().get("values", [])
    max_len = max((len(r) for r in values), default=0)
    grid = [row + [None] * (max_len - len(row)) for row in values]
    return _rows_from_grid(grid)


def decrypt_source_credentials(encrypted: str | None) -> dict:
    if not encrypted:
        return {}
    try:
        return json.loads(decrypt_secret(encrypted))
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# Postgres / Supabase external source — read-only, SSL required, only ever
# runs parameterized SELECTs against identifiers the caller fetched from
# information_schema itself (never free-text SQL from the user).

def _pg_connect(connection_string: str):
    import psycopg2

    if "sslmode=" not in connection_string:
        sep = "&" if "?" in connection_string else "?"
        connection_string = f"{connection_string}{sep}sslmode=require"
    return psycopg2.connect(connection_string, connect_timeout=10)


def postgres_list_tables(connection_string: str) -> list[str]:
    conn = _pg_connect(connection_string)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "select table_name from information_schema.tables "
                "where table_schema = 'public' and table_type = 'BASE TABLE' order by table_name"
            )
            return [r[0] for r in cur.fetchall()]
    finally:
        conn.close()


def postgres_read_table(connection_string: str, table_name: str, limit: int = 20000) -> tuple[list[str], list[dict]]:
    conn = _pg_connect(connection_string)
    try:
        with conn.cursor() as cur:
            valid_tables = set(postgres_list_tables(connection_string))
            if table_name not in valid_tables:
                raise ValueError("الجدول غير موجود أو غير مسموح بالوصول إليه")
            cur.execute(f'select * from "{table_name}" limit %s', (limit,))
            columns = [d[0] for d in cur.description]
            rows = [dict(zip(columns, row)) for row in cur.fetchall()]
            return columns, rows
    finally:
        conn.close()
